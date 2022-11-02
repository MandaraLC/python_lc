from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import Chrome
import random
import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import config
import json
import socket

class Flaginsertion:
    def __init__(self, filepath):
        self.issuccess = 0  # 是否执行成功，成功则退出循环
        self.filepath = filepath #excel储存路径
        self.sheetnames = ['新店插蓝旗', '新店插紫旗', '老店插蓝旗', '老店插紫旗']

    # 检测端口是否被占用
    def check_port_in_use(self, port, host='127.0.0.1'):
        s = None
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(0.5)
            s.connect((host, int(port)))
            return port
        except socket.error:
            return False
        finally:
            if s:
                s.close()

    # 查找元素是否存在 waittime等待的时间
    def findElement(self, driver, xpath, waittime):
        try:
            WebDriverWait(driver, waittime).until(EC.visibility_of_element_located((By.XPATH, xpath)))
            return True
        except:
            return False

    # 代理
    def ipadea_proxy(self):
        if config.ISUSERPROXY_SHOP == 1:
            # 加代理
            respone = requests.get(config.PROXY_URL)
            json0 = json.loads(respone.text)
            if json0['code'] == 0:
                return json0['data']
        else:
            # 不加代理
            return []

    # 写入文件
    def writefile(self, msg, cookie_file):
        fileInput = open(cookie_file, "a")
        fileInput.write(''.join(msg) + '\n')

    def broswer(self):
        #是否加代理
        if config.ISUSERPROXY_SHOP == 1:
            # 加代理
            respone = requests.get(config.PROXY_URL)
            json0 = json.loads(respone.text)
            if json0['code'] == 0:
                proxy = json0['data']
                for i in range(5):
                    try:
                        ipport = proxy[random.randint(0, len(proxy))]
                        break
                    except:
                        pass
            else:
                ipport = []
        else:
            # 不加代理
            ipport = []

        chromeOptions = webdriver.ChromeOptions()
        print("代理：", ipport)
        if ipport:
            chromeOptions.add_argument(f'--proxy-server=http://{ipport["ip"]}:{ipport["port"]}')

        #windows系统
        for i in range(10):
            port = str(random.randint(1, 65535))
            if self.check_port_in_use(port) == False:
                chromeOptions.add_argument('--remote-debugging-port=' + port)
                break
        chromeOptions.add_experimental_option('useAutomationExtension', False)
        chromeOptions.add_argument('--ignore-certificate-errors') #忽略证书错误
        # chromeOptions.add_experimental_option("excludeSwitches", ["enable-logging"])
        chromeOptions.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])
        chromeOptions.add_argument("--disable-blink-features=AutomationControlled")
        # chromeOptions.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片提升运行速度
        chromeOptions.add_argument("--start-maximized")
        chromeOptions.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36')
        driver = Chrome(executable_path="./chromedriver.exe", options=chromeOptions)
        # 绕过浏览器指纹
        # stealth(driver, languages=["en-US", "en"], vendor="Google Inc.", platform="Win32", webgl_vendor="Intel Inc.",
        #         renderer="Intel Iris OpenGL Engine", fix_hairline=True, )
        driver.set_page_load_timeout(60)  # 页面加载超时时间
        driver.set_script_timeout(60)  # 页面js加载超时时间
        driver.delete_all_cookies() #清空cookie
        return driver

    #登录淘宝
    def logintaobao(self, driver, content, shoptype):
        login_url = 'https://login.taobao.com/member/login.jhtml?spm=a21bo.jianhua.754894437.1.5af911d9kWNsIv&f=top&redirectURL=https%3A%2F%2Fwww.taobao.com%2F'
        driver.get(login_url)
        if (self.findElement(driver, "//input[@id='fm-login-id']", 10)):
            if shoptype == 1:
                account = config.TAOBAO_USERNAME_OLD
                password = config.TAOBAO_PASSWORD_OLD
            else:
                account = config.TAOBAO_USERNAME_NEW
                password = config.TAOBAO_PASSWORD_NEW
            # 账号
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-id']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-id']").send_keys(account)
            time.sleep(0.2)
            # 密码
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-password']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-password']").send_keys(password)
            print(f"已输入账号 / 密码：{account} / {password}")
            time.sleep(0.2)
            driver.find_element(by=By.XPATH, value="//button[@class='fm-button fm-submit password-login']").click()

            isloginsucc = 0
            if (self.findElement(driver, "//div[@class='indexwrap-2ymNW']",30) and driver.current_url.find("home.htm") > 0):
                print("登录成功！")
                isloginsucc = 1
            elif (self.findElement(driver, "//div[@id='login-error']",2)):
                print("账号或密码不正确！")
            elif (self.findElement(driver, "//input[@id='J_Phone_Checkcode']", 2)):
                # 需要手机验证登录
                verify = input("请输入验证码：")
                driver.find_element(by=By.XPATH, value="//input[@id='J_Phone_Checkcode']").send_keys(Keys.CONTROL + 'a')
                driver.find_element(by=By.XPATH, value="//input[@id='J_Phone_Checkcode']").send_keys(verify)
                time.sleep(1)
                driver.find_element(by=By.XPATH, value="//input[@class='ui-button ui-button-lorange']").click()
                isloginsucc == 1
            elif (self.findElement(driver, "//img[@id='bg-img']",2)):
                #需要滑块验证
                # ele = driver.find_element(by=By.XPATH, value="//span[@id='nc_1_n1z]")
                # # 实例化对象
                # action = ActionChains(driver)
                # # 拖动滑块
                # time.sleep(1)
                # action.drag_and_drop_by_offset(ele, xoffset=250, yoffset=0).perform()
                # time.sleep(1)
                isloginsucc == 1
            else:
                print("登录失败！")

            if isloginsucc == 1:
                for data in content:
                    if shoptype == 1:
                        if data['f'] == 4:
                            sheetname = self.sheetnames[2]
                        elif data['f'] == 5:
                            sheetname = self.sheetnames[3]
                    elif shoptype == 2:
                        if data['f'] == 4:
                            sheetname = self.sheetnames[0]
                        elif data['f'] == 5:
                            sheetname = self.sheetnames[1]
                    try:
                        #循环三次
                        for i in range(3):
                            self.issuccess = 0
                            self.dochaqi(data['o'], driver, data['f'], shoptype, sheetname)
                            if self.issuccess == 1:
                                break
                            if self.issuccess == 0 and i == 2:
                                self.excel(data['o'], data['f'], "插旗3次都失败", sheetname)
                    except Exception as e:
                        self.excel(data['o'], data['f'], "插旗失败，程序报错，已终止程序", sheetname)
                        print("程序报错：", e)
                        exit()
                    print("=========================================")
            else:
                print("登录失败")
        else:
            print("登录页面加载失败")

    #插旗
    def dochaqi(self, orderdata, driver, flag, shoptype, sheetname):
        if shoptype == 1:
            seller_id = '72317883'
        elif shoptype == 2:
            seller_id = '2200623874141'
        url = f'https://trade.taobao.com/trade/memo/update_sell_memo.htm?seller_id={seller_id}&biz_order_id={orderdata}&user_type=1&pageNum=1&auctionTitle=null&dateBegin=0&dateEnd=0&commentStatus=&buyerNick=&auctionStatus=PAID&logisticsService='
        driver.get(url)

        if (self.findElement(driver, "//table[@class='memo-editor']", 2)):
            driver.find_element(by=By.XPATH, value=f"//label[@for='flag{flag}']").click()
            time.sleep(0.2)
            text = driver.find_element(by=By.XPATH, value="//textarea[@id='memo']").text
            flag_text = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
            if text:
                discribe = f"{text}\n【程序插旗 有盒控需求 {flag_text}】"
            else:
                discribe = f"【程序插旗 有盒控需求 {flag_text}】"
            print(f"订单：{orderdata}\t备注：{discribe}")
            driver.find_element(by=By.XPATH, value="//textarea[@id='memo']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//textarea[@id='memo']").send_keys(discribe)
            time.sleep(0.2)
            #提交
            driver.execute_script('document.getElementById("form1").submit();')
            time.sleep(1)
            self.verify(orderdata, driver, flag, url, sheetname)
        else:
            print("页面加载失败！")

    #二次验证
    def verify(self, orderdata, driver, flag, url, sheetname):
        driver.get(url)
        isselected = driver.find_element(by=By.XPATH, value=f"//input[@id='flag{flag}']").is_selected()
        print("二次验证是否插旗成功：", isselected)
        if isselected == True:
            self.issuccess = 1
            self.excel(orderdata, flag, "成功", sheetname)
        else:
            print("插旗失败，正在重试！")

    #获取数据
    def getorderdata(self, excelpath):
        old_order = []
        new_order = []
        all_order = []
        excel = load_workbook(excelpath)
        all_sheet = excel.sheetnames
        for i in all_sheet:
            print(i)
            for column in excel[i].iter_cols():
                for cell2 in column:
                    # print(cell2.row, cell2.column, cell2.value)
                    # 获取第一列的数据，第一列的数据是蓝旗
                    if cell2.value is not None and cell2.row > 1 and cell2.column == 1:
                        if i == '老店':
                            old_order.append({'o': cell2.value.strip('="'), 'f': 4})
                        elif i == "新店":
                            new_order.append({'o': cell2.value.strip('="'), 'f': 4})
                    if cell2.value is not None and cell2.row > 1 and cell2.column == 2:
                        if i == '老店':
                            old_order.append({'o': cell2.value.strip('="'), 'f': 5})
                        elif i == "新店":
                            new_order.append({'o': cell2.value.strip('="'), 'f': 5})
        all_order.append(old_order)
        all_order.append(new_order)
        return all_order

    # 初始化excel
    def initexcel(self):
        excellist = ['订单编号', '插旗类型', '插旗状态']
        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        for key,value in enumerate(self.sheetnames):
            #创建sheet
            workbook.create_sheet(value, key)
            workbook[self.sheetnames[key]].append(excellist)
            workbook.save(filename=self.filepath)

    #执行成功时写入excel
    def excel(self, orderdata, flag, msg, sheetname):
        #写入excel文件
        infolist = []
        infolist.append(orderdata)
        if flag == 4:
            infolist.append("蓝旗")
        elif flag ==5:
            infolist.append("紫旗")
        infolist.append(msg)

        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        workbook[sheetname].append(infolist)
        workbook.save(filename=self.filepath)