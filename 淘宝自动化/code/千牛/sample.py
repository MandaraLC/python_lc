import subprocess
from time import sleep
from clicknium import clicknium as cc, locator, ui
from clicknium.common.enums import ClearHotKey,PreAction,InputTextBy,Location
from clicknium.common.models.mouselocation import MouseLocation
import os
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from pywinauto.keyboard import send_keys
import pyperclip

excel = load_workbook('./配置.xlsx')
all_sheet = excel.sheetnames
configdata = []
for i in all_sheet:
    for column in excel[i].iter_cols():
        for cell2 in column:
            if cell2.value is not None and cell2.row == 2 and cell2.column == 1:
                for a in excel[i][cell2.row]:
                    if a.value is not None:
                        configdata.append(a.value)

print("登录账号：", configdata[0], "\t登录密码：", configdata[1], "\t千牛启动程序路径：", configdata[2], "\t消息：", configdata[3])

def initexcel(filepath):
    excellist = ['千牛id', '状态']
    if os.path.exists(filepath):
        workbook = load_workbook(filename=filepath)
    else:
        workbook = Workbook()
    workbook.active.append(excellist)
    workbook.save(filename=filepath)

# 执行成功时写入excel
def excel(filepath, split, msg):
    # 写入excel文件
    infolist = []
    infolist.append(split[0])
    # infolist.append(split[1])
    infolist.append(msg)

    if os.path.exists(filepath):
        workbook = load_workbook(filename=filepath)
    else:
        workbook = Workbook()

    workbook.active.append(infolist)
    workbook.save(filename=filepath)

def get_name_list():
    '''
    千牛用户id和要发送的消息
    :return:
    '''
    datalist = []
    excel = load_workbook('./千牛id.xlsx')
    all_sheet = excel.sheetnames
    for i in all_sheet:
        for column in excel[i].iter_cols():
            for cell2 in column:
                if cell2.value is not None and cell2.row > 1 and cell2.column == 1:
                    rowvalues = []
                    for a in excel[i][cell2.row]:
                        if a.value is not None:
                            rowvalues.append(a.value)
                    datalist.append('\t'.join(rowvalues))
    return datalist

def start_process(path):
    '''
    启动千牛程序
    :param path:
    :return:
    '''
    # 通过 subprocess 库来启动千牛进程
    subprocess.Popen(path)
    sleep(5)
    print('千牛程序启动完成！')

def login(username, password):
    '''
    登录
    :param username:
    :param password:
    :return:
    '''
    # 首先清除掉已经有的账号
    ui(locator.aliworkbench.login.username).clear_text('send-hotkey',ClearHotKey.Home_ShiftEnd_Delete,PreAction.Click)
    # 填写用户名，这里无法使用UiElement对象的set_text 方法，它不是一个标准的edit输入框，我们可以使用全局的输入文本
    cc.send_text(username)

    # 发送enter快捷键，跳转到输入密码框
    cc.send_hotkey('{enter}')
    # 点击密码框
    ui(locator.aliworkbench.login.password).click()
    # 清空密码框
    ui(locator.aliworkbench.login.password).clear_text('send-hotkey',ClearHotKey.Home_ShiftEnd_Delete,PreAction.Click)
    # 输入密码
    cc.send_text(password)
    # 点击登录
    ui(locator.aliworkbench.login.btnLogin).click()
    print('登录完成！')

def send_msg(list_name, msg, filepath):
    # 循环list_name 列表，循环添加好友
    for data in list_name:
        try:
            userdata = data.split("\t")
            print("数据：", userdata)
            # 点击'搜索框'，并且清空文本，通过Home ShiftEnd Delete 组合快捷键来删除已有文本
            ui(locator.aliworkbench.im.txtSearch).clear_text('send-hotkey',ClearHotKey.Home_ShiftEnd_Delete,PreAction.Click)
            # '搜索框' 输入要添加的好友名称
            ui(locator.aliworkbench.im.txtSearch).set_text(userdata[0], InputTextBy.SendKeyAfterClick)
            # 点击'在网络中查找',这里该控件是没有控件树结构的，所以只能通过图像识别先识别位置，然后来点击它，由于各个电脑大小分辨率等不同，
            # 如果这个地方没有识别通过,你可以在你的电脑上重新使用录制器图像识别的方式抓取下该控件
            ui(locator.aliworkbench.im.btnSearchOnNetwork).click()
            # 睡眠1s,等待搜索结果出现
            sleep(0.5)
            # 这里需要点击第一条搜索结果，然后在聊天界面中点击添加好友按钮
            # 这里同样搜索结果是没有控件树结构的，并且每次内容都是不同的（也就不能使用图像识别方式来点击）
            # 所以我们通过定位'搜索框'+向下偏移40px坐标的方式 来点击第一条搜索结果
            ui(locator.aliworkbench.im.txtSearch).click(mouse_location=MouseLocation(Location.Center,yoffset=40))

            pyperclip.copy(msg)
            sleep(0.3)
            send_keys("^v")

            # cc.send_text(msg)
            sleep(0.3)
            # ui(locator.aliworkbench.btn_send).click()
            cc.send_hotkey('{enter}')
            sleep(0.3)
            cc.send_hotkey('{esc}')
            print("发送成功！")
            excel(filepath, userdata, '发送成功')
        except Exception as e:
            print("发送失败！", e)
            excel(filepath, userdata, '失败')
        print("========================================")

def main():
    # 获取千牛用户id和要发送的消息
    list_name = get_name_list()
    # 启动千牛进程
    start_process(configdata[2])
    # 登录
    login(configdata[0], configdata[1])
    sleep(4)
    # 点击 '接待中心 '
    ui(locator.aliworkbench.btnIM).click()
    # 等待 '接待中心' 界面出现
    cc.wait_appear(locator.aliworkbench.im.imWindow)
    # 循环发送消息
    filepath = f'./excel/千牛发送消息{time.strftime("%Y%m%d%H%M%S", time.localtime(time.time()))}.xlsx'
    #初始化
    initexcel(filepath)
    #发送消息
    send_msg(list_name, configdata[3], filepath)
    sleep(1)

if __name__ == "__main__":
    main()
