from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import Chrome
import random
import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import config
import json
import socket
from selenium_stealth import stealth

class Taobao:
    def __init__(self, filepath):
        self.issuccess = 0  # 是否执行成功，成功则退出循环
        self.filepath = filepath #excel储存路径

    # 检测端口是否被占用
    def check_port_in_use(self, port, host='127.0.0.1'):
        s = None
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(0.5)
            s.connect((host, int(port)))
            return port
        except socket.error:
            return False
        finally:
            if s:
                s.close()

    # 查找元素是否存在 waittime等待的时间
    def findElement(self, driver, xpath, waittime):
        try:
            WebDriverWait(driver, waittime).until(EC.visibility_of_element_located((By.XPATH, xpath)))
            return True
        except:
            return False

    # 代理
    def ipadea_proxy(self):
        if config.ISUSERPROXY_SHOP == 1:
            # 加代理
            respone = requests.get(config.PROXY_URL)
            json0 = json.loads(respone.text)
            if json0['code'] == 0:
                return json0['data']
        else:
            # 不加代理
            return []

    # 写入文件
    def writefile(self, msg, cookie_file):
        fileInput = open(cookie_file, "a")
        fileInput.write(''.join(msg) + '\n')

    def broswer(self):
        '''
        #是否加代理
        if config.ISUSERPROXY_SHOP == 1:
            # 加代理
            respone = requests.get(config.PROXY_URL)
            json0 = json.loads(respone.text)
            if json0['code'] == 0:
                proxy = json0['data']
                for i in range(5):
                    try:
                        ipport = proxy[random.randint(0, len(proxy))]
                        break
                    except:
                        pass
            else:
                ipport = []
        else:
            # 不加代理
            ipport = []

        chromeOptions = webdriver.ChromeOptions()
        print("代理：", ipport)
        if ipport:
            chromeOptions.add_argument(f'--proxy-server=http://{ipport["ip"]}:{ipport["port"]}')

        #windows系统
        for i in range(10):
            port = str(random.randint(1, 65535))
            if self.check_port_in_use(port) == False:
                chromeOptions.add_argument('--remote-debugging-port=' + port)
                break
        chromeOptions.add_experimental_option('useAutomationExtension', False)
        chromeOptions.add_argument('--ignore-certificate-errors') #忽略证书错误
        # chromeOptions.add_experimental_option("excludeSwitches", ["enable-logging"])
        chromeOptions.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])
        chromeOptions.add_argument("--disable-blink-features=AutomationControlled")
        # chromeOptions.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片提升运行速度
        chromeOptions.add_argument("--start-maximized")
        chromeOptions.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36')
        driver = Chrome(executable_path="../chromedriver.exe", options=chromeOptions)
        # 绕过浏览器指纹
        stealth(driver, languages=["en-US", "en"], vendor="Google Inc.", platform="Win32", webgl_vendor="Intel Inc.",
                renderer="Intel Iris OpenGL Engine", fix_hairline=True, )
        driver.set_page_load_timeout(60)  # 页面加载超时时间
        driver.set_script_timeout(60)  # 页面js加载超时时间
        driver.delete_all_cookies() #清空cookie
        '''
        # 这是手动启动浏览器的代码
        # os.system('chrome.exe --remote-debugging-port=9222 --user-data-dir="C:/Users/29047/Desktop/chrome"')
        chromeOptions = webdriver.ChromeOptions()
        chromeOptions.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        chromeOptions.add_argument('disable-infobars')
        # chromeOptions.add_argument("window-size=1920,1080")
        # chromeOptions.add_experimental_option('useAutomationExtension', False)
        chromeOptions.add_argument(
            'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36')
        driver = webdriver.Chrome(executable_path="../chromedriver.exe", options=chromeOptions)

        return driver

    #登录淘宝
    def logintaobao(self, driver, content, username, password):
        login_url = 'https://login.taobao.com/member/login.jhtml?f=top&redirectURL=https%3A%2F%2Fwww.taobao.com%2F'
        driver.get(login_url)
        if (self.findElement(driver, "//input[@id='fm-login-id']", 10)):
            # 账号
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-id']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-id']").send_keys(username)
            time.sleep(0.2)
            # 密码
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-password']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-password']").send_keys(password)
            print(f"已输入账号 / 密码：{username} / {password}")
            time.sleep(0.2)
            driver.find_element(by=By.XPATH, value="//button[@class='fm-button fm-submit password-login']").click()

            isloginsucc = 0
            if (self.findElement(driver, "//div[@class='indexwrap-2ymNW']",30) and driver.current_url.find("home.htm") > 0):
                print("登录成功！")
                isloginsucc = 1
            elif (self.findElement(driver, "//div[@id='login-error']",2)):
                print("账号或密码不正确！")
            elif (self.findElement(driver, "//input[@id='J_Phone_Checkcode']", 2)):
                # 需要手机验证登录
                verify = input("请输入验证码：")
                driver.find_element(by=By.XPATH, value="//input[@id='J_Phone_Checkcode']").send_keys(Keys.CONTROL + 'a')
                driver.find_element(by=By.XPATH, value="//input[@id='J_Phone_Checkcode']").send_keys(verify)
                time.sleep(1)
                driver.find_element(by=By.XPATH, value="//input[@class='ui-button ui-button-lorange']").click()
                isloginsucc == 1
            elif (self.findElement(driver, "//img[@id='bg-img']",2)):
                #需要滑块验证
                # ele = driver.find_element(by=By.XPATH, value="//span[@id='nc_1_n1z]")
                # # 实例化对象
                # action = ActionChains(driver)
                # # 拖动滑块
                # time.sleep(1)
                # action.drag_and_drop_by_offset(ele, xoffset=250, yoffset=0).perform()
                # time.sleep(1)
                isloginsucc == 1
            else:
                print("登录失败！")

            if isloginsucc == 1:
                for data in content:
                    slipt = data.strip().split("\t")
                    newslipt = []
                    newslipt.append(slipt[0].strip('="'))
                    newslipt.append(slipt[1].strip('="'))
                    print("数据：", newslipt)
                    # try:
                    #循环三次
                    for i in range(3):
                        self.issuccess = 0
                        self.suresend(newslipt, driver)
                        if self.issuccess == 1:
                            break
                        if self.issuccess == 0 and i == 2:
                            self.excel(newslipt, "发货三次都失败")
                    # except Exception as e:
                    #     self.excel(slipt, "发货失败，程序报错，已终止程序")
                    #     print("程序报错：", e)
                    #     exit()
                    print("=========================================")
                    time.sleep(3)
        else:
            print("登录页面加载失败")

    def suresend(self, slipt, driver):
        driver.get("https://www.baidu.com")
        time.sleep(1)
        url = f'https://trade.taobao.com/trade/detail/trade_order_detail.htm?biz_order_id={slipt[1]}&sifg=1'
        driver.get(url)

        if (self.findElement(driver, "//h3[text()='当前订单状态：交易关闭']", 2) or self.findElement(driver, "//h3[text()='当前订单状态：商品已拍下，等待买家付款']", 2)):
            print("交易关闭或买家未付款")
            self.suresend0(slipt, driver)
        # elif(self.findElement(driver, "//a[text()='修改收货地址']", 1)):
        #
        #     driver.get(f"https://trade.taobao.com/trade/modifyDeliverAddress.htm?bizOrderId={slipt[1]}")
        #     if(self.findElement(driver, "//i[@class='next-icon next-icon-eye-close next-small']", 3)):
        #         try:
        #             driver.find_element(by=By.XPATH, value="//i[@class='next-icon next-icon-eye-close next-small']").click()
        #             time.sleep(0.5)
        #             info = []
        #             # em = driver.find_element(by=By.XPATH, value="//div[@role='gridcell']").text
        #             # print(em)
        #             username = driver.find_element(by=By.XPATH, value="//input[@placeholder='输入新收货人']").get_attribute("value")
        #             phone = driver.find_element(by=By.XPATH, value="//input[@placeholder='输入新收货电话']").get_attribute("value")
        #             addr = driver.find_element(by=By.XPATH, value="//input[@data-spm-anchor-id='0.0.0.i1.26c11a49grVp5q']").get_attribute("value")
        #             info.append(username)
        #             info.append(phone)
        #             info.append(addr)
        #             self.issuccess = 1
        #             # self.excelinfo(info, slipt[1])
        #         except Exception as e:
        #             print("获取数据报错（或者出现验证码）：", e)
        #     else:
        #         print("找不到眼镜按钮")
        #     driver.switch_to.default_content()
        elif (self.findElement(driver, "//ul[@class='tabs-mod__tabs-nav___1dAjY']", 1)):
            driver.find_element(by=By.XPATH, value="//ul[@class='tabs-mod__tabs-nav___1dAjY']/li[2]/a").click()
            text = driver.find_element(by=By.XPATH, value="//div[@class='logistics-panel-mod__line-info___2l8PA']/span[2]/span").text
            if text and text.find("*") < 0:
                self.issuccess = 1
                if text.find("，") >= 0:
                    splitinfo = text.split("，")
                else:
                    splitinfo = text.split(",")
                print("info：", splitinfo)
                self.excelinfo(splitinfo, slipt[1])
            else:
                # print("显示星号...")
                # self.suresend0(slipt, driver)
                # time.sleep(60 * 10)
                print("显示星号，去旧版获取...")
                # 先去获取用户信息
                getinfourl = f'https://wuliu.taobao.com/user/order_detail_new.htm?old=1&trade_id={slipt[1]}'
                driver.get(getinfourl)
                if (self.findElement(driver, "//img[@id='eyeAddress']", 2)):
                    info = ''
                    for i in range(5):
                        comfirmdel = driver.find_element(By.XPATH, "//img[@id='eyeAddress']")
                        driver.execute_script("arguments[0].click();", comfirmdel)
                        time.sleep(2)
                        info = driver.find_element(By.XPATH, "//span[@id='receiverInfo']").text
                        print("用户信息：", info)
                        if info.find("*") >= 0:
                            print("有星号")
                            time.sleep(2)
                        else:
                            break

                    if info.find("*") >= 0:
                        print("旧版也显示星号，暂停30S....")
                        time.sleep(30)
                    else:
                        splitinfo = info.split(",")
                        print("成功：", splitinfo)
                        self.excelinfo2(splitinfo, slipt[1])
                        self.issuccess = 1
        else:
            print("页面加载失败")
            self.suresend0(slipt, driver)

    def suresend0(self, slipt, driver):
        #此链接判断订单状态：已发货、未发货、交易关闭
        driver.get("https://www.baidu.com")
        time.sleep(1)
        url = f'https://wuliu2.taobao.com/user/consign2.htm?tradeId={slipt[1]}&x-frames-allow-from=wuliu2'
        driver.get(url)
        if (self.findElement(driver, "//div[@class='package-logistic-info']", 1)):
            if (self.findElement(driver, "//i[@class='qn_iconfont cursor-pointer']", 2)):
                try:
                    driver.find_element(by=By.XPATH, value="//i[@class='qn_iconfont cursor-pointer']").click()
                    time.sleep(1)
                    info = driver.find_element(by=By.XPATH, value="//i[@class='qn_iconfont cursor-pointer']/../span").text

                    if info.find("*") < 0:
                        splitinfo = info.split(",")
                        print("info：", splitinfo)
                        self.excelinfo1(splitinfo, slipt[1])
                        self.issuccess = 1
                    else:
                        print("也显示星号，去旧版获取...")
                        # 先去获取用户信息
                        getinfourl = f'https://wuliu.taobao.com/user/order_detail_new.htm?old=1&trade_id={slipt[1]}'
                        driver.get(getinfourl)
                        if (self.findElement(driver, "//img[@id='eyeAddress']", 2)):
                            info = ''
                            for i in range(5):
                                comfirmdel = driver.find_element(By.XPATH, "//img[@id='eyeAddress']")
                                driver.execute_script("arguments[0].click();", comfirmdel)
                                time.sleep(2)
                                info = driver.find_element(By.XPATH, "//span[@id='receiverInfo']").text
                                print("用户信息：", info)
                                if info.find("*") >= 0:
                                    print("有星号")
                                    time.sleep(2)
                                else:
                                    break

                            if info.find("*") >= 0:
                                print("旧版也显示星号，暂停30S....")
                                time.sleep(30)
                            else:
                                splitinfo = info.split(",")
                                print("成功：", splitinfo)
                                self.excelinfo2(splitinfo, slipt[1])
                                self.issuccess = 1
                except Exception as e:
                    self.suresend(slipt, driver)
                    print("报错4：", e)

        elif (self.findElement(driver, "//span[@aria-haspopup='true']", 1)):
            print("此订单未发货")
            self.issuccess = 1
        else:
            print("页面加载失败，正在重试")

    # 初始化excel
    def initexcel(self):
        excellist = ['姓名', '电话', '地址', '订单号']
        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()
        workbook.active.append(excellist)
        workbook.save(filename=self.filepath)

    #执行成功时写入excel
    def excel(self, split, msg):
        #写入excel文件
        infolist = []
        infolist.append(split[0])
        infolist.append(split[1])
        infolist.append(msg)

        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        workbook.active.append(infolist)
        workbook.save(filename=self.filepath)

    def excelinfo(self, splitinfo, order):
        # 写入excel文件
        infolist = []
        # infolist.append(splitinfo[1])
        # infolist.append(splitinfo[2])
        # infolist.append(splitinfo[0])
        # infolist.append(order)
        infolist.append(splitinfo[0].strip())
        if str(splitinfo[1]).find("86-") >= 0:
            infolist.append(splitinfo[1].replace('86-', '').strip())
        else:
            infolist.append(splitinfo[1].strip())
        infolist.append(splitinfo[2].strip())

        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        workbook.active.append(infolist)
        workbook.save(filename=self.filepath)

    def excelinfo1(self, splitinfo, order):
        # 写入excel文件
        infolist = []
        infolist.append(splitinfo[1].strip())
        infolist.append(splitinfo[2].strip())
        infolist.append(splitinfo[0].strip())

        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        workbook.active.append(infolist)
        workbook.save(filename=self.filepath)


    def excelinfo2(self, splitinfo, order):
        # 写入excel文件
        infolist = []
        infolist.append(splitinfo[2].strip())
        infolist.append(splitinfo[3].strip())
        infolist.append(splitinfo[0].strip())
        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        workbook.active.append(infolist)
        workbook.save(filename=self.filepath)