from taobao import Taobao
import time
from openpyxl import load_workbook
import config

old_order = []
new_order = []
# excel = load_workbook('./data/order_data.xlsx')
excel = load_workbook('./data/order_data_20220810.xlsx')
all_sheet = excel.sheetnames
print(all_sheet)
for i in all_sheet:
    for column in excel[i].iter_cols():
        for cell2 in column:
            if cell2.value is not None and cell2.row > 1 and cell2.column == 1:
                if i == '老店':
                    rowvalues = []
                    for a in excel[i][cell2.row]:
                        if a.value is not None:
                            rowvalues.append(str(a.value))
                    old_order.append('\t'.join(rowvalues))
                elif i == "新店":
                    rowvalues = []
                    for a in excel[i][cell2.row]:
                        if a.value is not None:
                            rowvalues.append(a.value)
                    new_order.append('\t'.join(rowvalues))

if old_order:
    # 实例化
    filepath = f'./excel/淘宝info_老店.xlsx'
    taobao = Taobao(filepath=filepath)
    driver = taobao.broswer()
    #初始化excel
    #taobao.initexcel()
    #登录淘宝
    taobao.logintaobao(driver, old_order, config.TAOBAO_USERNAME_OLD, config.TAOBAO_PASSWORD_OLD)
    try:
        driver.quit()
    except:
        pass
else:
    print("老店没有订单数据！")

# if new_order:
#     # 实例化
#     filepath = f'./excel/淘宝info_新店.xlsx'
#     taobao = Taobao(filepath=filepath)
#     driver = taobao.broswer()
#     #初始化excel
#     #taobao.initexcel()
#     #登录淘宝
#     taobao.logintaobao(driver, new_order, config.TAOBAO_USERNAME_NEW, config.TAOBAO_PASSWORD_NEW)
#     try:
#         driver.quit()
#     except:
#         pass
# else:
#     print("新店没有订单数据！")

#新2010
#老1952