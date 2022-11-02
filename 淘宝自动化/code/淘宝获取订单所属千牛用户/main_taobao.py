from taobao import Taobao
import time
from openpyxl import load_workbook

old_order = []
new_order = []
excel = load_workbook('./data/订单号.xlsx')
all_sheet = excel.sheetnames
for i in all_sheet:
    for column in excel[i].iter_cols():
        for cell2 in column:
            if cell2.value is not None and cell2.row > 1 and cell2.column == 1:
                if i == '老店':
                    rowvalues = []
                    for a in excel[i][cell2.row]:
                        if a.value is not None:
                            rowvalues.append(a.value)
                    old_order.append('\t'.join(rowvalues))
                elif i == "新店":
                    rowvalues = []
                    for a in excel[i][cell2.row]:
                        if a.value is not None:
                            rowvalues.append(a.value)
                    new_order.append('\t'.join(rowvalues))

#获取账号密码
oldaccinfo = []
newaccinfo = []
excel = load_workbook('./data/账号密码.xlsx')
all_sheet = excel.sheetnames
for i in all_sheet:
    for column in excel[i].iter_cols():
        for cell2 in column:
            if cell2.value is not None and cell2.row == 2 and cell2.column == 1:
                if i == '老店':
                    for a in excel[i][cell2.row]:
                        if a.value is not None:
                            oldaccinfo.append(a.value)
                elif i == "新店":
                    for a in excel[i][cell2.row]:
                        if a.value is not None:
                            newaccinfo.append(a.value)

if len(oldaccinfo) == 0 or len(newaccinfo) == 0:
    print("请先配置账号密码！")
    exit()

if old_order:
    # 实例化
    filepath = f'./excel/获取订单所属千牛用户_老店_{time.strftime("%Y%m%d%H%M%S", time.localtime(time.time()))}.xlsx'
    taobao = Taobao(filepath=filepath)
    driver = taobao.broswer()
    #初始化excel
    taobao.initexcel()
    #登录淘宝
    taobao.logintaobao(driver, old_order, oldaccinfo[0], oldaccinfo[1])
    try:
        driver.quit()
    except:
        pass
else:
    print("老店没有订单数据！")

if new_order:
    # 实例化
    filepath = f'./excel/获取订单所属千牛用户_新店_{time.strftime("%Y%m%d%H%M%S", time.localtime(time.time()))}.xlsx'
    taobao = Taobao(filepath=filepath)
    driver = taobao.broswer()
    #初始化excel
    taobao.initexcel()
    #登录淘宝
    taobao.logintaobao(driver, new_order, newaccinfo[0], newaccinfo[1])
    try:
        driver.quit()
    except:
        pass
else:
    print("新店没有订单数据！")