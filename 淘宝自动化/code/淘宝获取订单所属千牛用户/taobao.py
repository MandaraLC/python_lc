import re

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import Chrome
import random
import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from selenium_stealth import stealth

import config
import json
import socket
from selenium.webdriver.common.action_chains import ActionChains


class Taobao:
    def __init__(self, filepath):
        self.issuccess = 0  # 是否执行成功，成功则退出循环
        self.filepath = filepath  # excel储存路径

    # 检测端口是否被占用
    def check_port_in_use(self, port, host='127.0.0.1'):
        s = None
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(0.5)
            s.connect((host, int(port)))
            return port
        except socket.error:
            return False
        finally:
            if s:
                s.close()

    # 查找元素是否存在 waittime等待的时间
    def findElement(self, driver, xpath, waittime):
        try:
            WebDriverWait(driver, waittime).until(EC.visibility_of_element_located((By.XPATH, xpath)))
            return True
        except:
            return False

    # 代理
    def ipadea_proxy(self):
        if config.ISUSERPROXY_SHOP == 1:
            # 加代理
            respone = requests.get(config.PROXY_URL)
            json0 = json.loads(respone.text)
            if json0['code'] == 0:
                return json0['data']
        else:
            # 不加代理
            return []

    # 写入文件
    def writefile(self, msg, cookie_file):
        fileInput = open(cookie_file, "a")
        fileInput.write(''.join(msg) + '\n')

    def broswer(self):
        #是否加代理
        # if config.ISUSERPROXY_SHOP == 1:
        #     # 加代理
        #     respone = requests.get(config.PROXY_URL)
        #     json0 = json.loads(respone.text)
        #     if json0['code'] == 0:
        #         proxy = json0['data']
        #         for i in range(5):
        #             try:
        #                 ipport = proxy[random.randint(0, len(proxy))]
        #                 break
        #             except:
        #                 pass
        #     else:
        #         ipport = []
        # else:
        #     # 不加代理
        #     ipport = []
        #
        #
        # chromeOptions = webdriver.ChromeOptions()
        # print("代理：", ipport)
        # if ipport:
        #     chromeOptions.add_argument(f'--proxy-server=http://{ipport["ip"]}:{ipport["port"]}')
        #
        # #windows系统
        # for i in range(10):
        #     port = str(random.randint(1, 65535))
        #     if self.check_port_in_use(port) == False:
        #         chromeOptions.add_argument('--remote-debugging-port=' + port)
        #         break
        # # chromeOptions.add_argument("--headless")
        # chromeOptions.add_experimental_option('useAutomationExtension', False)
        # chromeOptions.add_argument('--ignore-certificate-errors') #忽略证书错误
        # # chromeOptions.add_experimental_option("excludeSwitches", ["enable-logging"])
        # chromeOptions.add_experimental_option('excludeSwitches', ['enable-automation','enable-logging'])
        # chromeOptions.add_argument("--disable-blink-features=AutomationControlled")
        # # chromeOptions.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片提升运行速度
        # chromeOptions.add_argument("--start-maximized")
        # chromeOptions.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36')
        # driver = Chrome(executable_path="./chromedriver.exe", options=chromeOptions)
        # # 绕过浏览器指纹
        # stealth(driver, languages=["en-US", "en"], vendor="Google Inc.", platform="Win32", webgl_vendor="Intel Inc.",
        #         renderer="Intel Iris OpenGL Engine", fix_hairline=True, )
        # # 输入 stealth.min.js 文件路径
        # with open('./stealth.min.js') as f:
        #     js = f.read()
        # driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        #     "source": js
        # })
        # driver.set_page_load_timeout(60)  # 页面加载超时时间
        # driver.set_script_timeout(60)  # 页面js加载超时时间
        # driver.delete_all_cookies() #清空cookie

        # # 这是手动启动浏览器的代码
        # os.system('chrome.exe --remote-debugging-port=9222 --user-data-dir="D:/chrome"')
        chromeOptions = webdriver.ChromeOptions()
        chromeOptions.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        chromeOptions.add_argument('disable-infobars')
        # chromeOptions.add_argument("window-size=1920,1080")
        # chromeOptions.add_experimental_option('useAutomationExtension', False)
        chromeOptions.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36')
        driver = webdriver.Chrome(options=chromeOptions)
        return driver

    # 登录淘宝
    def logintaobao(self, driver, content, username, password):
        login_url = 'https://login.taobao.com/member/login.jhtml?f=top&redirectURL=https%3A%2F%2Fwww.taobao.com%2F'
        driver.get(login_url)
        if (self.findElement(driver, "//input[@id='fm-login-id']", 10)):
            # 账号
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-id']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-id']").send_keys(username)
            time.sleep(0.2)
            # 密码
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-password']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-password']").send_keys(password)
            print(f"已输入账号 / 密码：{username} / {password}")

            time.sleep(0.2)
            isloginsucc = 0
            try:
                driver.find_element(by=By.XPATH, value="//button[@class='fm-button fm-submit password-login']").click()
            except:
                print("登录失败！")
            if (self.findElement(driver, "//div[@class='indexwrap-2ymNW']", 120) and driver.current_url.find("home.htm") > 0):
                print("登录成功！")
                isloginsucc = 1
            elif (self.findElement(driver, "//div[@id='login-error']", 2)):
                print("账号或密码不正确！")
            elif (self.findElement(driver, "//input[@id='J_Phone_Checkcode']", 2)):
                # 需要手机验证登录
                verify = input("请输入验证码：")
                driver.find_element(by=By.XPATH, value="//input[@id='J_Phone_Checkcode']").send_keys(Keys.CONTROL + 'a')
                driver.find_element(by=By.XPATH, value="//input[@id='J_Phone_Checkcode']").send_keys(verify)
                time.sleep(1)
                driver.find_element(by=By.XPATH, value="//input[@class='ui-button ui-button-lorange']").click()
                isloginsucc == 1
            elif (self.findElement(driver, "//img[@id='bg-img']", 2)):
                # 需要滑块验证
                # ele = driver.find_element(by=By.XPATH, value="//span[@id='nc_1_n1z]")
                # # 实例化对象
                # action = ActionChains(driver)
                # # 拖动滑块
                # time.sleep(1)
                # action.drag_and_drop_by_offset(ele, xoffset=250, yoffset=0).perform()
                # time.sleep(1)
                isloginsucc == 1
            else:
                print("登录失败！")

            if isloginsucc == 1:
                for data in content:
                    data0 = data.strip('="')
                    print("数据：", data0)
                    # 循环三次
                    # for i in range(3):
                    try:
                        self.issuccess = 0
                        self.get_order(data0, driver)
                        if self.issuccess == 1:
                            break
                    except Exception as e:
                        self.excel('', '', '程序报错')
                        print("程序报错：", e)
                    print("=========================================")
        else:
            print("登录页面加载失败")

    # 获取订单
    def get_order(self, data, driver):
        driver.get("https://trade.taobao.com/trade/itemlist/list_sold_items.htm?action=itemlist/SoldQueryAction&event_submit_do_query=1&auctionStatus=PAID&tabCode=waitSend")
        # 获取到cookie
        sel_cookies = driver.get_cookies()  # 获取selenium侧的cookies
        jar = requests.cookies.RequestsCookieJar()  # 先构建RequestsCookieJar对象
        for i in sel_cookies:
            jar.set(i['name'], i['value'], domain=i['domain'], path=i['path'])
        session = requests.session()  # requests以session会话形式访问网站
        session.cookies.update(jar)  # 将配置好的RequestsCookieJar对象加入到requests形式的session会话中

        print(f"正在请求订单数据...")
        # 请求接口
        api_url = 'https://trade.taobao.com/trade/itemlist/asyncSold.htm?event_submit_do_query=1&_input_charset=utf8&sifg=0'
        # post数据
        postdata = {
            "auctionType": "0",
            "close": "0",
            "pageNum": "1",
            "pageSize": "15",
            "payDateBegin": "0",
            "payDateEnd": "0",
            "queryMore": "false",
            "rxAuditFlag": "0",
            "rxElectronicAllFlag": "0",
            "rxElectronicAuditFlag": "0",
            "rxHasSendFlag": "0",
            "rxOldFlag": "0",
            "rxSendFlag": "0",
            "rxSuccessflag": "0",
            "rxWaitSendflag": "0",
            "sellerMemo": "0",
            "tradeTag": "0",
            "useCheckcode": "false",
            "useOrderInfo": "false",
            "errorCheckcode": "false",
            "action": "itemlist/SoldQueryAction",
            "orderId": str(data),
            "prePageNo": "1",
            "sifg": "",
            "queryLabelValues": '[{"label":"订单编号","value":"'+str(data)+'","index":7},{"label":"买家昵称","value":"","index":4},{"label":"宝贝名称","value":"","index":2},{"label":"商品ID","value":"","index":1}]',
        }
        # 请求头
        header = {
            "accept": "application/json, text/javascript, */*; q=0.01",
            "accept-encoding": "gzip, deflate, br",
            "accept-language": "zh-CN,zh;q=0.9",
            "bx-v": "2.2.2",
            "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
            "origin": "https://trade.taobao.com",
            "referer": "https://trade.taobao.com/trade/itemlist/list_sold_items.htm",
            "sec-ch-ua-mobile": "?0",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.60 Safari/537.36",
            "x-requested-with": "XMLHttpRequest"
        }
        respone = session.post(url=api_url, data=postdata, headers=header)
        if respone.status_code == 200:
            respone_data = json.loads(respone.text)
            for item in respone_data['mainOrders']:
                try:
                    print("订单号：", item['id'], '\t千牛用户ID：', item['buyer']['decodeNick'])
                    self.excel(item['id'], item['buyer']['decodeNick'], '成功')
                except:
                    decodeNick = re.findall('userID=(.*)&sign=', item['buyer']['memberUrl'])
                    print("订单号：", item['id'], '\t千牛用户ID：', decodeNick[0].strip())
                    self.excel(item['id'], decodeNick[0].strip(), '成功')
        else:
            self.excel('', '', '接口请求失败')
            print("接口请求失败：", respone.text)
        time.sleep(1)

    # 初始化excel
    def initexcel(self):
        excellist = ['订单编号', '千牛用户ID', '状态']
        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()
        workbook.active.append(excellist)
        workbook.save(filename=self.filepath)

    # 执行成功时写入excel
    def excel(self, id, nick, msg):
        # 写入excel文件
        infolist = []
        infolist.append(id)
        infolist.append(nick)
        infolist.append(msg)

        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        workbook.active.append(infolist)
        workbook.save(filename=self.filepath)
