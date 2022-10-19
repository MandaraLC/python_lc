from openpyxl import load_workbook
from refund import Refund
import time

excel = load_workbook('./data/配置.xlsx')
all_sheet0 = excel.sheetnames
configdata = []
for i in all_sheet0:
    for column in excel[i].iter_cols():
        for cell2 in column:
            if cell2.value is not None and cell2.row == 2 and cell2.column == 1:
                for a in excel[i][cell2.row]:
                    if a.value is not None:
                        configdata.append(a.value)

try:
    print("程序开始运行...")
    starttime = time.time()
    print("程序运行开始时间：", starttime)
    #实例化
    filename = f'退款{time.strftime("%Y%m%d%H%M%S", time.localtime(time.time()))}.xlsx'
    filepath = f'./excel/{filename}'
    refundobj = Refund(filepath=filepath)

    all_order = [] #所有预定单号
    #要搜索的文件
    excel = load_workbook(f'data/{configdata[0]}')
    all_sheet = excel.sheetnames
    for i in all_sheet:
        for column in excel[i].iter_cols():
            for cell2 in column:
                if cell2.value is not None and cell2.column == 2:
                    all_order.append(cell2.value)

    print(f"共{len(all_order)}个订单号：", all_order)

    if len(all_order) > 0:
        #初始化
        refundobj.initexcel()
        print("开始去定金表搜索...")
        #开始去定金表搜索预订单号
        refundobj.find_false_in_xlsx(f'./data/{configdata[1]}', all_order)

    endtime = time.time()
    print("程序运行结束时间：", endtime)
    print(f"程序运行总耗时：{endtime-starttime}秒")
    print("程序运行完毕！请在excel目录下查看运行结果")
except Exception as e:
    print(e)