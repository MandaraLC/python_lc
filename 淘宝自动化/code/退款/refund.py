import os
from openpyxl import Workbook
from openpyxl import load_workbook

class Refund:
    #实例化
    def __init__(self, filepath):
        self.filepath = filepath  # excel储存路径

    def find_false_in_sheet(self, sheet, sheetname, orderarr):
        for column in sheet.iter_cols():
            for cell2 in column:
                if cell2.value is not None:
                    for order in orderarr:
                        info2 = str(cell2.value).find(order)
                        if info2 >= 0 and cell2.column == 2:
                            print(f"sheet名称：{sheetname}", f"\t行数：{cell2.row}", f"\t列数：{cell2.column}")
                            rowvalues = []
                            for i in sheet[cell2.row]:
                                rowvalues.append(i.value)
                            print(f"该列数据：{rowvalues}")
                            self.excel(rowvalues) #写入excel
                            print("-------------------------------------")

    def find_false_in_xlsx(self, file_name, orderarr):
        wb = load_workbook(file_name)
        all_sheets = wb.sheetnames
        print("所有sheet名称：", all_sheets)
        for sheetname in range(len(all_sheets)):
            print("当前搜索sheet名称：", all_sheets[sheetname])
            self.find_false_in_sheet(wb[all_sheets[sheetname]], all_sheets[sheetname], orderarr)
            print("================================================")

    # 初始化excel
    def initexcel(self):
        excellist = ['退款日期', '预定单号', '商品名称', 'ID', "数量", "总金额", "退款类型", '店名', '支付宝账号',
                     '姓名', '登记人', '发售日', '备注']
        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()
        workbook.active.append(excellist)
        workbook.save(filename=self.filepath)

    #写入excel
    def excel(self, rowvalues):
        # 写入excel文件
        infolist = []
        infolist.append('')
        infolist.append(rowvalues[1])
        infolist.append(rowvalues[3])
        infolist.append(rowvalues[7])
        infolist.append(rowvalues[5])
        infolist.append(rowvalues[4])
        infolist.append('')
        infolist.append('')
        infolist.append('')
        infolist.append('')
        infolist.append('')
        infolist.append(rowvalues[11])
        infolist.append(rowvalues[13])

        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()
        workbook.active.append(infolist)
        workbook.save(filename=self.filepath)

