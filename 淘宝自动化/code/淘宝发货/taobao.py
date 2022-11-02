from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import Chrome
import random
import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from selenium_stealth import stealth

import config
import json
import socket
from selenium.webdriver.common.action_chains import ActionChains


class Taobao:
    def __init__(self, filepath):
        self.issuccess = 0  # 是否执行成功，成功则退出循环
        self.filepath = filepath  # excel储存路径

    # 检测端口是否被占用
    def check_port_in_use(self, port, host='127.0.0.1'):
        s = None
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(0.5)
            s.connect((host, int(port)))
            return port
        except socket.error:
            return False
        finally:
            if s:
                s.close()

    # 查找元素是否存在 waittime等待的时间
    def findElement(self, driver, xpath, waittime):
        try:
            WebDriverWait(driver, waittime).until(EC.visibility_of_element_located((By.XPATH, xpath)))
            return True
        except:
            return False

    # 代理
    def ipadea_proxy(self):
        if config.ISUSERPROXY_SHOP == 1:
            # 加代理
            respone = requests.get(config.PROXY_URL)
            json0 = json.loads(respone.text)
            if json0['code'] == 0:
                return json0['data']
        else:
            # 不加代理
            return []

    # 写入文件
    def writefile(self, msg, cookie_file):
        fileInput = open(cookie_file, "a")
        fileInput.write(''.join(msg) + '\n')

    def broswer(self):
        #是否加代理
        # if config.ISUSERPROXY_SHOP == 1:
        #     # 加代理
        #     respone = requests.get(config.PROXY_URL)
        #     json0 = json.loads(respone.text)
        #     if json0['code'] == 0:
        #         proxy = json0['data']
        #         for i in range(5):
        #             try:
        #                 ipport = proxy[random.randint(0, len(proxy))]
        #                 break
        #             except:
        #                 pass
        #     else:
        #         ipport = []
        # else:
        #     # 不加代理
        #     ipport = []
        #
        #
        # chromeOptions = webdriver.ChromeOptions()
        # print("代理：", ipport)
        # if ipport:
        #     chromeOptions.add_argument(f'--proxy-server=http://{ipport["ip"]}:{ipport["port"]}')
        #
        # #windows系统
        # for i in range(10):
        #     port = str(random.randint(1, 65535))
        #     if self.check_port_in_use(port) == False:
        #         chromeOptions.add_argument('--remote-debugging-port=' + port)
        #         break
        # # chromeOptions.add_argument("--headless")
        # chromeOptions.add_experimental_option('useAutomationExtension', False)
        # chromeOptions.add_argument('--ignore-certificate-errors') #忽略证书错误
        # # chromeOptions.add_experimental_option("excludeSwitches", ["enable-logging"])
        # chromeOptions.add_experimental_option('excludeSwitches', ['enable-automation','enable-logging'])
        # chromeOptions.add_argument("--disable-blink-features=AutomationControlled")
        # # chromeOptions.add_argument('blink-settings=imagesEnabled=false')  # 不加载图片提升运行速度
        # chromeOptions.add_argument("--start-maximized")
        # chromeOptions.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36')
        # driver = Chrome(executable_path="./chromedriver.exe", options=chromeOptions)
        # # 绕过浏览器指纹
        # stealth(driver, languages=["en-US", "en"], vendor="Google Inc.", platform="Win32", webgl_vendor="Intel Inc.",
        #         renderer="Intel Iris OpenGL Engine", fix_hairline=True, )
        # # 输入 stealth.min.js 文件路径
        # with open('./stealth.min.js') as f:
        #     js = f.read()
        # driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        #     "source": js
        # })
        # driver.set_page_load_timeout(60)  # 页面加载超时时间
        # driver.set_script_timeout(60)  # 页面js加载超时时间
        # driver.delete_all_cookies() #清空cookie

        # # 这是手动启动浏览器的代码
        # os.system('chrome.exe --remote-debugging-port=9222 --user-data-dir="D:/chrome"')
        chromeOptions = webdriver.ChromeOptions()
        chromeOptions.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        chromeOptions.add_argument('disable-infobars')
        # chromeOptions.add_argument("window-size=1920,1080")
        # chromeOptions.add_experimental_option('useAutomationExtension', False)
        chromeOptions.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36')
        driver = webdriver.Chrome(options=chromeOptions)

        return driver

    def get_track(self, distance):
        # 移动轨迹
        track = []

        sum = 10
        for i in range(30):
            sum += i
            if sum < distance:
                track.append(sum)
            else:
                track.append(sum - distance)
                break
        print(track)
        return track

    def get_track0(self, distance):  # distance为传入的总距离
        # 移动轨迹
        track = []
        # 当前位移
        current = 0
        # 减速阈值
        mid = distance * 4 / 5
        # 计算间隔
        t = 0.05
        # 初速度
        v = 3

        while current < distance:
            if current < mid:
                # 加速度为2
                a = 10
            else:
                # 加速度为-2
                a = -3
            v0 = v
            # 当前速度
            v = v0 + a * t
            # 移动距离
            move = v0 * t + 1 / 2 * a * t * t
            # 当前位移
            current += move
            # 加入轨迹
            track.append(round(move))

        return track

    def check_stock(self, driver):
        nc_1_n1z = driver.find_element(by=By.XPATH, value="//span[@id='nc_1_n1z']")
        tracks = self.get_track(400)
        ActionChains(driver).click_and_hold(nc_1_n1z).perform()
        try:
            for x in tracks:
                ActionChains(driver).move_by_offset(xoffset=x, yoffset=0).perform()
        except:
            pass
        time.sleep(0.5)
        ActionChains(driver).release().perform()

    # 登录淘宝
    def logintaobao(self, driver, content, username, password):
        login_url = 'https://login.taobao.com/member/login.jhtml?f=top&redirectURL=https%3A%2F%2Fwww.taobao.com%2F'
        driver.get(login_url)
        if (self.findElement(driver, "//input[@id='fm-login-id']", 10)):
            # 账号
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-id']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-id']").send_keys(username)
            time.sleep(0.2)
            # 密码
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-password']").send_keys(Keys.CONTROL + 'a')
            driver.find_element(by=By.XPATH, value="//input[@id='fm-login-password']").send_keys(password)
            print(f"已输入账号 / 密码：{username} / {password}")

            time.sleep(0.2)
            isloginsucc = 0
            # self.slide(driver)
            try:
                driver.find_element(by=By.XPATH, value="//button[@class='fm-button fm-submit password-login']").click()
            except:
                print("登录失败！")
            if (self.findElement(driver, "//div[@class='indexwrap-2ymNW']", 120) and driver.current_url.find(
                    "home.htm") > 0):
                print("登录成功！")
                isloginsucc = 1
            elif (self.findElement(driver, "//div[@id='login-error']", 2)):
                print("账号或密码不正确！")
            elif (self.findElement(driver, "//input[@id='J_Phone_Checkcode']", 2)):
                # 需要手机验证登录
                verify = input("请输入验证码：")
                driver.find_element(by=By.XPATH, value="//input[@id='J_Phone_Checkcode']").send_keys(Keys.CONTROL + 'a')
                driver.find_element(by=By.XPATH, value="//input[@id='J_Phone_Checkcode']").send_keys(verify)
                time.sleep(1)
                driver.find_element(by=By.XPATH, value="//input[@class='ui-button ui-button-lorange']").click()
                isloginsucc == 1
            elif (self.findElement(driver, "//img[@id='bg-img']", 2)):
                # 需要滑块验证
                # ele = driver.find_element(by=By.XPATH, value="//span[@id='nc_1_n1z]")
                # # 实例化对象
                # action = ActionChains(driver)
                # # 拖动滑块
                # time.sleep(1)
                # action.drag_and_drop_by_offset(ele, xoffset=250, yoffset=0).perform()
                # time.sleep(1)
                isloginsucc == 1
            else:
                print("登录失败！")

            if isloginsucc == 1:
                for data in content:
                    slipt = data.strip().split("\t")
                    newslipt = []
                    newslipt.append(slipt[0].strip('="'))
                    newslipt.append(slipt[1].strip('="'))
                    print("数据：", newslipt)
                    # try:
                    # 循环三次
                    for i in range(3):
                        try:
                            self.issuccess = 0
                            self.suresend(newslipt, driver)
                            if self.issuccess == 1:
                                break
                        except Exception as e:
                            print("程序报错：", e)
                        if self.issuccess == 0 and i == 2:
                            self.excel(newslipt, '', "发货三次都失败")
                    # except Exception as e:
                    #     self.excel(slipt, "发货失败，程序报错，已终止程序")
                    #     print("程序报错：", e)
                    #     exit()
                    print("=========================================")
        else:
            print("登录页面加载失败")

    # 出现滑块
    def slide(self, driver):
        ishaveslide = 0
        if self.findElement(driver, '//iframe[@id="baxia-dialog-content"]', 1):
            ishaveslide = 1

        if (self.findElement(driver, '//span[@id="nc_1_n1z"]', 1) or ishaveslide == 1):
            print("出现滑块")
            if 1 == ishaveslide:
                driver.switch_to.frame('baxia-dialog-content')
            # time.sleep(500)
            self.check_stock(driver)
            # # 定位到滑块按钮元素
            # ele_button = driver.find_element(By.XPATH, '//span[@id="nc_1_n1z"]')
            # # 定位到滑块区域元素
            # ele = driver.find_element(By.XPATH, '//div[@id="nc_1__scale_text"]')
            #
            # if 1 == ishaveslide:
            #     driver.switch_to.frame('baxia-dialog-content')
            #
            # ActionChains(driver).drag_and_drop_by_offset(ele_button, ele.size['width'], ele.size['height']).perform()
            if 1 == ishaveslide:
                driver.switch_to.default_content()
        else:
            print("没有出现滑块")

    def suresend(self, slipt, driver):
        # 此链接判断订单状态：已发货、未发货、交易关闭
        url_status = f'https://trade.taobao.com/trade/detail/trade_order_detail.htm?biz_order_id={slipt[1]}'
        driver.get(url_status)
        # self.slide(driver)
        if (self.findElement(driver, "//h3[text()='当前订单状态：交易关闭']", 2)):
            self.excel(slipt, '', "用户已退款或交易关闭")
            print("用户已退款或交易关闭！")
            self.issuccess = 1
        elif (self.findElement(driver, "//h3[text()='当前订单状态：商品已拍下，等待买家付款']", 1)):
            self.issuccess = 1
            self.excel(slipt, '', "商品已拍下，等待买家付款")
            print("商品已拍下，等待买家付款！")
        else:
            info = ''
            # # 先去获取用户信息
            # getinfourl = f'https://wuliu.taobao.com/user/order_detail_new.htm?old=1&trade_id={slipt[1]}'
            # driver.get(getinfourl)
            # if (self.findElement(driver, "//img[@id='eyeAddress']", 2)):
            #     for i in range(5):
            #         # 点击未显示的元素
            #         comfirmdel = driver.find_element(By.XPATH, "//img[@id='eyeAddress']")
            #         driver.execute_script("arguments[0].click();", comfirmdel)
            #         # driver.find_element(By.XPATH, "//img[@id='eyeAddress']").click()
            #         info = driver.find_element(By.XPATH, "//span[@id='receiverInfo']").text
            #         print("用户信息：", info)
            #         if info.find("*") >= 0:
            #             time.sleep(1)
            #         else:
            #             break
            #
            # time.sleep(0.5)
            url = f'https://wuliu2.taobao.com/user/consign2.htm?tradeId={slipt[1]}&x-frames-allow-from=wuliu2'
            driver.get(url)
            # if info.find("*") >= 0:
            #     try:
            #         if (self.findElement(driver, "//i[@class='qn_iconfont cursor-pointer']", 1)):
            #             driver.find_element(By.XPATH, "//i[@class='qn_iconfont cursor-pointer']").click()
            #             time.sleep(0.5)
            #             info = driver.find_element(By.XPATH, "//i[@class='qn_iconfont cursor-pointer']/../span").text
            #             print("用户信息：", info)
            #     except Exception as e:
            #         pass

            if (self.findElement(driver, "//div[@class='cp-code-selector']", 2)):
                # 点击未显示的元素
                comfirmdel = driver.find_element(By.XPATH, "//div[@class='cp-code-selector']")
                driver.execute_script("arguments[0].click();", comfirmdel)
                # 输入物流单号
                driver.find_element(by=By.XPATH, value="//span[@class='next-select-trigger-search']/input").send_keys(
                    Keys.CONTROL + 'a')
                driver.find_element(by=By.XPATH, value="//span[@class='next-select-trigger-search']/input").send_keys(
                    slipt[0])
                if (self.findElement(driver, "//li[@class='next-menu-footer']/div/div[2]/div[1]/span[2]", 10)):
                    try:
                        # 选择其他物流公司
                        comfirmdel1 = driver.find_element(By.XPATH, "//li[@class='next-menu-footer']/div/div[2]/div[1]/span[2]")
                        driver.execute_script("arguments[0].click();", comfirmdel1)
                        # 选择其他物流公司
                        # driver.find_element(by=By.XPATH, value="//li[@class='next-menu-footer']/div/div[2]/div[1]/span[2]").click()
                        time.sleep(0.2)
                        # 确认并发货
                        omfirmdel0 = driver.find_element(By.XPATH, "//button[@class='next-btn next-medium next-btn-primary']")
                        driver.execute_script("arguments[0].click();", omfirmdel0)
                        # driver.find_element(by=By.XPATH, value="//button[@class='next-btn next-medium next-btn-primary']").click()
                    except Exception as e:
                        print("发货失败！888")
                        pass
                    # 二次验证是否发货成功
                    self.issendsucc(slipt, info, driver)
                else:
                    print("发货失败，无法选择物流公司")
                    # self.excel(slipt, "发货失败，无法选择物流公司")
            elif (self.findElement(driver, "//div[@class='package-logistic-info']", 1)):
                self.issuccess = 1
                # self.excel(slipt, "此订单已发货，无法重复发货")
                if self.findElement(driver, f"//span[text()='{slipt[0]}']", 2):
                    print("发货成功666")
                    self.excel(slipt, info, "发货成功")
                else:
                    print("此订单已发货")
                    self.excel(slipt, info, "此订单已发货")
            elif (self.findElement(driver, "//div[text()='订单不存在']", 1)):
                print("没有此订单")
                self.issuccess = 1
                self.excel(slipt, '', "发货失败，没有此订单")
            elif (self.findElement(driver, "//div[@id='err']", 1) and driver.current_url.find("error.html") > 0):
                print("没有此订单")
                self.issuccess = 1
                self.excel(slipt, '', "发货失败，没有此订单")
            else:
                print("页面加载失败，正在重试")

    # 是否发货成功
    def issendsucc(self, slipt, info, driver):
        # 判断是否出现滑块
        # self.slide(driver)
        if (self.findElement(driver, "//div[@class='operation-container']", 10) and driver.current_url.find(
                "batch_consign.htm") > 0):
            print("正在二次验证...")
            time.sleep(0.5)
            url = f'https://wuliu2.taobao.com/user/consign2.htm?tradeId={slipt[1]}&x-frames-allow-from=wuliu2'
            driver.get(url)
            if (self.findElement(driver, "//div[@class='package-logistic-info']", 2)):
                print("发货成功")
                self.issuccess = 1
                self.excel(slipt, info, "发货成功")
            else:
                # 判断是否出现滑块
                print("发货失败，正在重新发货")
        else:
            print("发货失败！")
        time.sleep(2)

    # 初始化excel
    def initexcel(self):
        # excellist = ['物流单号', '订单编号', '发货信息', '发货状态']
        excellist = ['物流单号', '订单编号', '发货状态']
        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()
        workbook.active.append(excellist)
        workbook.save(filename=self.filepath)

    # 执行成功时写入excel
    def excel(self, split, info, msg):
        # 写入excel文件
        infolist = []
        infolist.append(split[0])
        infolist.append(split[1])
        # infolist.append(info)
        infolist.append(msg)

        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        workbook.active.append(infolist)
        workbook.save(filename=self.filepath)
