import time
import requests
import json
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from lxml import etree

class Cz89:
    #初始化
    def __init__(self):
        self.headers = {
            "accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            "accept-encoding":"gzip, deflate, br",
            "accept-language":"zh-CN,zh;q=0.9",
            "cache-control":"max-age=0",
            "referer":"https://m.cz89.com/",
            "sec-ch-ua-mobile":"?0",
            "sec-fetch-dest":"document",
            "sec-fetch-mode":"navigate",
            "sec-fetch-site":"same-origin",
            "sec-fetch-user":"?1",
            "upgrade-insecure-requests":"1",
            "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.60 Safari/537.36"
        }
        ycqihao = self.getycqihao()
        filename = f'yc{ycqihao}_数据{time.strftime("%Y-%m-%d", time.localtime(time.time()))}.txt'
        self.filepath = f'./data/{filename}'
        try:
            os.remove(self.filepath)
        except:
            pass

    def join_list(self, item):
        """处理列表到字符串"""
        return ",".join(item)

    #获取所预测的期号
    def getycqihao(self):
        header = {
            "Host": "datachart.500.com",
            "Content-Type": "text/html; charset=gb2312",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3861.400 QQBrowser/10.7.4313.400",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Language": "zh-CN,zh;q=0.9"
        }
        # ssq
        url1 = 'https://datachart.500.com/ssq/history/history.shtml'

        respone1 = requests.get(url=url1, headers=header)
        html = etree.HTML(respone1.text)
        qihao = self.join_list(html.xpath("//input[@id='end']/@value"))
        return int(qihao)+1

    #获取每个预测种类的数据
    def getdatanums(self):
        new_data = [] #生成的新数组
        qihao = str(self.getycqihao())
        print("预测期号：", qihao)
        for page in range(10):
            get_data_api = f'https://m.cz89.com/ssq/item_8.htm?tdsourcetag=s_pcqq_aiomsg&p={page+1}'
            respone_get_data = requests.get(url=get_data_api, headers=self.headers)
            html = etree.HTML(respone_get_data.text)

            items = html.xpath("//ul[@class='commenList']/li")
            # items.pop(0)
            for item in items:
                text = self.join_list(item.xpath("./a/text()"))
                href = self.join_list(item.xpath("./a/@href"))
                if text.find(qihao[-2:]) >= 0:
                    # print(text, href)
                    new_data.append({
                        'href':href,
                        'text':text
                    })
            # if len(new_data) == 0:
            #     break
        print(f"共{len(new_data)}条数据 ==>", new_data)
        self.getdedaildata(new_data)

    #进入详情页获取数据
    def getdedaildata(self, new_data):
        print(f"共{len(new_data)}条数据，数据获取完成！以下是访问详情页面，并将数据写excel中...")
        for data in new_data:
            detail_page_url = f"https://m.cz89.com{data['href']}"
            print("当前访问：", data['text'], detail_page_url)

            respone_detail_data = requests.get(url=detail_page_url, headers=self.headers)
            html = etree.HTML(respone_detail_data.text)

            items = html.xpath("//article[@class='article']/p")
            fileInput = open(self.filepath, "a")
            fileInput.write(f"当前访问：{data['text']}\t{detail_page_url}\n")

            print(f"共{len(items)}个p元素")
            for i in range(len(items)):
                # print(f"第{i+1}个元素")
                item = self.join_list(html.xpath(f"//article[@class='article']/p[{i+1}]/text()"))
                time0 = self.join_list(html.xpath(f"//article[@class='article']/p[{i+1}]/time/text()"))
                if item:
                    if time0:
                        fileInput.write(f"{item}{time0}\n")
                    else:
                        fileInput.write(f"{item}\n")
            #将数据写入文件
            fileInput.write(f"===================================\n")
            print("============================")

    #写入文件
    def excel(self, name, expertname, value):
        # 写入excel文件
        infolist = []
        infolist.append(expertname)
        infolist.append(value)
        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()
        workbook[name].append(infolist)
        workbook.save(filename=self.filepath)

    # 初始化excel
    def initexcel(self):
        # 表头
        tatal = ['昵称', '数据']
        if os.path.exists(self.filepath):
            workbook = load_workbook(filename=self.filepath)
        else:
            workbook = Workbook()

        for key, value in enumerate(self.sheetnames):
            # 创建sheet
            workbook.create_sheet(value, key)
            workbook[self.sheetnames[key]].append(tatal)
            workbook.save(filename=self.filepath)