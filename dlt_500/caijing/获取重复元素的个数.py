from openpyxl import load_workbook
import os
excel = load_workbook(f'excel/数据2022-07-19.xlsx')
writefilepath = "./data/重复个数2022-07-19.txt"
try:
    os.remove(writefilepath)
except:
    pass
all_sheet = excel.sheetnames
for i in all_sheet:
    old_order = []
    for column in excel[i].iter_cols():
        for cell2 in column:

            if cell2.value is not None and cell2.row > 1 and cell2.column == 2:
                #print(cell2.row, cell2.column, cell2.value)
                if cell2.value.find(",") >= 0:
                    split0 = cell2.value.split(",")
                    for a in split0:
                        old_order.append(a)
                else:
                    old_order.append(cell2.value)

    myset = set(old_order)
    datajson = {}
    for item in myset:
        datajson[item] = old_order.count(item)

    # 降序排序
    desc_data = sorted(datajson.items(), key=lambda x: x[1], reverse=True)
    if desc_data:
        print(i, "：", desc_data)
        fileInput = open(writefilepath, "a")
        fileInput.write(f"{i}：{str(desc_data)}\n")